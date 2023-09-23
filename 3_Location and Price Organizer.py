# Import the openpyxl library for xlsx files management
import openpyxl

# Open the file where you have the data to manipulate
wb = openpyxl.load_workbook('file_path')
# Open the desired worksheet in the xlsx file
sheet = wb.active

# Initialize a counter on the second row, skipping the table headers
i=2

# Write the header for the state column
sheet.cell(row = 1, column = 7).value = 'State'

# Loop through the rows
while i < sheet.max_row+1 :

    # Acess the price cell in the i row, B column
    price = sheet.cell(row = i, column = 2)

    # Acess the cell in the i row, C column
    location = sheet.cell(row = i, column = 3)

    # Acess the cell in the i row, G column
    state = sheet.cell(row = i, column = 7)

    # Check if there is no price
    if (price.value == "Contactez l'annonceur" or location.value == "Contactez l'annonceur"):
        
        # Delete the row that doesn't contain the price
        sheet.delete_rows(i)
    
    # Check if the price and location are swapped
    if (location.value[0] in ['0','1','2','3','4','5','6','7','8','9']) :
        location.value = price.value
        price.value = location.value

    try :
        state.value = location.value.split(',')[1]
        location.value = location.value.split(',')[0]
    except :
        print('An exception has occurred')

    # Remove the currency and spaces in the price and convert it to an int
    try :
        price.value = int(price.value.split('DT')[0].replace(" ",""))
    except :
        print('An exception has occurred')

    print(state.value,location.value,price.value)

    i+=1

# Save the output workbooks
print('-------------------------Saving file-------------------------')
wb.save('file_path')