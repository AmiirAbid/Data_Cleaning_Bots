# Import the openpyxl library for xlsx files management
import openpyxl

# Open the file where you have the data to manipulate
wb = openpyxl.load_workbook('file_path')

# Open the desired worksheet in the xlsx file
sheet = wb['sheet_name']

# Define the function that extracts the surface from the details
def surface_extractor(surface):

    # Initialize a new string
    s = ''

    # Iterate through each character in reverse order
    for char in reversed(surface):

        # Check if the character is numeric or a comma
        if char.isnumeric() or char == ',':

            # Prepend the character to the new string (to maintain the correct order)
            s = char + s

        else:

            # Stop the loop when a non-numeric character is encountered
            break

    # Return the new string
    return s

# Loop through the rows, starting from the second row (skipping headers)
for i in range(2, sheet.max_row + 1):
    
    print(sheet.cell(row=i, column=1).value)

    # Access the cells in the i-th row
    surface_habitable = sheet.cell(row=i, column=4)
    surface_terrain = sheet.cell(row=i, column=5)
    details = sheet.cell(row=i, column=6)

    # Check if there is no "Surface Habitable" and there is in the description
    if (surface_habitable.value is None or "Surf" not in surface_habitable.value) and (details.value is not None and 'Surf habitable' in details.value):
        
        # Split the details into 2 strings with "m² Surf habitable" as a separator
        surface = details.value.split("m² Surf habitable")[0]
        surface_habitable.value = surface

    # Check if there is no "Surface Terrain" and there is in the description    
    if (surface_terrain.value is None or "Surf" not in surface_terrain.value) and (details.value is not None and 'Surf terrain' in details.value):
        
        # Split the details into 2 strings with "m² Surf terrain" as a separator
        surface = details.value.split("m² Surf terrain")[0]
        surface_terrain.value = surface

    if surface_habitable.value is not None and 'm²' in surface_habitable.value:
        
        # Extract the exact surface without any alphabetic characters
        surface_habitable.value = surface_extractor(surface_habitable.value.split('m²')[0])
        print(surface_habitable.value)
        
        # Modify the type of the "Surface Habitable" to float
        try :
            surface_habitable.value = float(surface_habitable.value)
        except :
            print('An exception has occurred')
    

    elif surface_habitable.value is not None and 'm²' not in surface_habitable.value:

        # Extract the exact surface without any alphabetic characters
        surface_habitable.value = surface_extractor(surface_habitable.value)

    if surface_terrain.value is not None and 'm²' in surface_terrain.value:

        # Extract the exact surface without any alphabetic characters
        surface_terrain.value = surface_extractor(surface_terrain.value.split('m²')[0])
        print(surface_terrain.value)
        try :
            surface_terrain.value = float(surface_terrain.value)
        except :
            print('An exception has occurred')

    elif surface_terrain.value is not None and 'm²' not in surface_terrain.value:

        # Extract the exact surface without any alphabetic characters
        surface_terrain.value = surface_extractor(surface_terrain.value)

    # Check if the "surface habitable" and "surface terrain" are swapped
    if (type(surface_habitable.value) == float and type(surface_terrain.value) == float) and surface_terrain.value < surface_habitable.value:
        
        # Swap the values
        surface_habitable.value, surface_terrain.value = surface_terrain.value, surface_habitable.value
        print(surface_habitable.value,'/',surface_terrain.value)

# Save the output workbooks
print('-------------------------Saving files-------------------------')
wb.save('file_path')
