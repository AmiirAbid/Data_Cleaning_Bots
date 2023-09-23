# Import the openpyxl library for xlsx files management
import openpyxl

# Open the file where you have the data to manipulate
wb = openpyxl.load_workbook('file_path')
# Open the desired worksheet in the xlsx file
sheet = wb['Sheet']

# Create a new workbook for the specific data
wb1 = openpyxl.Workbook()
ws1 = wb1.active

# Make a wordlist for the wanted keywords
l1 = ['''write the keywords''']
l2 = ['''write the keywords''']

# Make a wordlist for the unwanted keywords
l3 = ['''write the keywords''']

# Make the header for each column
for j in range(1,8) :
    ws1.cell(row = 1, column = j).value = sheet.cell(row = 1, column = j).value

# Initialize a counter on the second row, skipping the table headers
i=2

# Loop through the rows
while i < sheet.max_row+1 :
    
    # Acess the cell in the i row, A column
    cell = sheet.cell(row = i, column = 1)

    # Check if the cell contains a keyword from the land wordlist
    if (cell.value != None) and any(keyword in cell.value for keyword in l1) and any(keyword in cell.value for keyword in l2) and all(keyword not in cell.value for keyword in l3) :

        # Add this row to the lands file
        ws1.append([cell.value for cell in sheet[i]])

    i+=1

# Save the output workbooks
print('-------------------------Saving files-------------------------')
wb1.save('file_path')