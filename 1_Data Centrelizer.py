import os
import openpyxl

# Define the folder containing XLSX files
folder_path = 'file path'

# Initialize an empty list to store all rows
all_rows = []

# Function to process each XLSX file in the folder
def process_xlsx_file(file_path):
    try:
        # Open the XLSX file
        wb = openpyxl.load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Iterate through rows and append to all_rows
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                all_rows.append(row)

    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")

# Loop through all files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)
        process_xlsx_file(file_path)

# Create a new workbook and worksheet for the combined data
combined_wb = openpyxl.Workbook()
combined_sheet = combined_wb.active

# Write all the rows from all_files into the combined worksheet
for row in all_rows:
    combined_sheet.append(row)

# Save the combined data to a new XLSX file
combined_file_path = os.path.join(folder_path, 'combined_data.xlsx')
combined_wb.save(combined_file_path)

print(f"Combined data saved to {combined_file_path}")
