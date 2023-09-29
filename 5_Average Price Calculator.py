import openpyxl

# Load the workbook and sheet
wb = openpyxl.load_workbook('file_path')
sheet = wb['sheet_name']

# Dictionaries to keep track of total price per m² and count for each state
state_totals = {}
state_counts = {}

# Dictionaries to keep track of total price per m² and count for each city
city_totals = {}
city_counts = {}

# Iterate over the rows
for row in range(2, sheet.max_row + 1):
    try:
        # Get price value from column B, remove spaces, and convert to int
        price = sheet.cell(row=row, column=2).value
        
        # Get state from column G
        state = sheet.cell(row=row, column=7).value

        # Get city from column C
        city = sheet.cell(row=row, column=3).value + ',' + state

        if price is not None :

            # Update the totals and counts for the state
            state_totals[state] = state_totals.get(state, 0) + price
            state_counts[state] = state_counts.get(state, 0) + 1

            # Update the totals and counts for the city
            city_totals[city] = city_totals.get(city, 0) + price
            city_counts[city] = city_counts.get(city, 0) + 1

    except Exception as e:
        print(f"Error in row {row}: {e}")

# Create a new workbook for results
result_wb = openpyxl.Workbook()

# Create new sheets for states and cities
states_sheet = result_wb.create_sheet(title="Average Price in States")
cities_sheet = result_wb.create_sheet(title="Average Price in Cities")

# Write headers for states
states_sheet['A1'] = "State"
states_sheet['B1'] = "Average Price"
states_sheet['C1'] = "Number of listings"

# Write headers for cities
cities_sheet['A1'] = "City"
cities_sheet['B1'] = "State"
cities_sheet['C1'] = "Average Price"
cities_sheet['D1'] = "Number of listings"

# Calculate and write average price for each state
row_num = 2
for state, total in state_totals.items():
    avg_price = total / state_counts[state]
    states_sheet[f'A{row_num}'] = state
    states_sheet[f'B{row_num}'] = avg_price
    states_sheet[f'C{row_num}'] = state_counts[state]
    row_num += 1

# Calculate and write average price for each city
row_num = 2
for city, total in city_totals.items():
    avg_price = total / city_counts[city]
    city_info = city.split(',')
    cities_sheet[f'A{row_num}'] = city_info[0]  # City name
    cities_sheet[f'B{row_num}'] = city_info[1]  # State
    cities_sheet[f'C{row_num}'] = avg_price
    cities_sheet[f'D{row_num}'] = city_counts[city]
    row_num += 1

# Save the results to a new file
result_wb.save('file_path')

# Close both workbooks
wb.close()
result_wb.close()