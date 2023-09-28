import openpyxl
from collections import defaultdict


# Load the Excel workbook
file_path = 'file_path'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Define the list of keywords for property types
keywords = ['''write the keywords''']  # Update this list as required

# Dictionary to store the counts
count_dict = defaultdict(lambda: defaultdict(int))

# Loop through the rows and count
for row in sheet.iter_rows(min_row=2, values_only=True):  # assuming 1st row is header
    title = row[0]
    city = row[2]
    state = row[6]
    
    for keyword in keywords:
        if title is not None and keyword in title.lower():
            count_dict[(state, city)][keyword] += 1

# Create a new workbook to store the results
output_workbook = openpyxl.Workbook()
new_sheet = output_workbook.active
new_sheet.title = "Listings"

# Write header
header = ["State", "City"] + keywords
new_sheet.append(header)

for (state, city), keyword_counts in count_dict.items():
    row = [state, city] + [keyword_counts[keyword] for keyword in keywords]
    new_sheet.append(row)

# Save the workbook with the counts
output_workbook.save('file_path')