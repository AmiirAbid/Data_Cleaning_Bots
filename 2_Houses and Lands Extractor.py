# Import the openpyxl library for xlsx files management
import openpyxl

# Open the file where you have the data to manipulate
wb = openpyxl.load_workbook('file_path')

# Open the desired worksheet in the xlsx file
sheet = wb['sheet_name']

# Create a new workbook for Lands
wb1 = openpyxl.Workbook()
ws1 = wb1.active

# Create a new workbook for Houses
wb2 = openpyxl.Workbook()
ws2 = wb2.active

# Make a wordlist for keywords that indicates that the property type is a land
l = ['''write the keywords''']
    
# Make a wordlist for keywords that indicates that the property type is a house
h = ['''write the keywords''']

# add the header for each column
for j in range(1,8) :
    ws1.cell(row = 1, column = j).value = ws2.cell(row = 1, column = j).value = sheet.cell(row = 1, column = j).value

# Initialize a counter on the second row, skipping the table headers
i=2

# Loop through the rows
while i < sheet.max_row+1 :
    
    # Acess the title cell in the i row, A column
    title = sheet.cell(row = i, column = 1)
    
    # Check if the property is a land
    if (title.value != None) and any(keyword in title.value for keyword in l) :

        # Print the accessed cell
        print(i,'LAND: ',title.value)

        # Add this row to the lands file
        ws1.append([title.value for title in sheet[i]])

    # Check if the property is a house
    elif (title.value != None) and any(keyword in title.value for keyword in h) :
        
        # Print the accessed cell
        print(i,'HOUSE: ',title.value)

        # Add this row to the houses file
        ws2.append([title.value for title in sheet[i]])

    i+=1

# Save the output workbooks
print('-------------------------Saving files-------------------------')
wb1.save('file_path')
wb2.save('file_path')